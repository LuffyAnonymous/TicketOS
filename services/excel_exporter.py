import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Excel file paths at the project root
EXCEL_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "liveticketgroup_customers.xlsx"
)

EXCEL_HISTORY_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "liveticketgroup_customer_history.xlsx"
)

def split_customer_name(full_name):
    if not full_name or str(full_name).strip() in ("", "-", "None"):
        return "", ""
    parts = str(full_name).strip().split(maxsplit=1)
    if len(parts) == 2:
        return parts[0], parts[1]
    elif len(parts) == 1:
        return parts[0], ""
    return "", ""

def format_purchase_datetime(sale_date):
    if not sale_date:
        return "N/A", "N/A"
    from datetime import datetime
    if isinstance(sale_date, str):
        from core.helpers import parse_sale_datetime
        dt = parse_sale_datetime(sale_date)
        if not dt:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ"):
                try:
                    dt = datetime.strptime(sale_date, fmt)
                    break
                except:
                    pass
        if dt:
            sale_date = dt
        else:
            parts = sale_date.split()
            if len(parts) == 2:
                return parts[0], parts[1]
            return sale_date, "N/A"
            
    if hasattr(sale_date, "strftime"):
        return sale_date.strftime("%Y-%m-%d"), sale_date.strftime("%H:%M:%S")
        
    return "N/A", "N/A"

def extract_order_from_details(details):
    if not details:
        return {}
        
    # Handle pre-split names if present
    first_name = details.get("first_name") or details.get("First Name")
    last_name = details.get("last_name") or details.get("Last Name")
    if first_name is not None or last_name is not None:
        first_name = str(first_name or "").strip()
        last_name = str(last_name or "").strip()
    else:
        full_name = details.get("billing_full_name") or details.get("customer_name") or ""
        first_name, last_name = split_customer_name(full_name)
        
    # Handle email
    email = details.get("email") or details.get("Email") or ""
    email = str(email).strip()
    if not email or email.lower() == "none":
        email = "N/A"
        
    # Handle mobile number
    mobile = details.get("mobile_number") or details.get("Mobile Number") or details.get("billing_mobile") or details.get("phone_number") or details.get("Phone Number") or ""
    mobile = str(mobile).strip()
    
    # Handle game purchased
    game = details.get("game_purchased") or details.get("Game Purchased") or details.get("event_name") or details.get("event") or ""
    game = str(game).strip()
    
    # Handle purchase date & time
    sale_date = details.get("sale_date") or details.get("Sale Date")
    purchase_date, purchase_time = format_purchase_datetime(sale_date)
    
    if purchase_date == "N/A" and ("purchase_date" in details or "Purchase Date" in details):
        purchase_date = details.get("purchase_date") or details.get("Purchase Date")
    if purchase_time == "N/A" and ("purchase_time" in details or "Purchase Time" in details):
        purchase_time = details.get("purchase_time") or details.get("Purchase Time")
        
    # Handle quantity
    qty = details.get("quantity") or details.get("Quantity") or details.get("Total Tickets Bought") or details.get("total_tickets_bought")
    try:
        qty = int(qty) if qty is not None else 1
    except:
        qty = 1
    if qty <= 0:
        qty = 1
        
    return {
        "first_name": first_name,
        "last_name": last_name,
        "mobile_number": mobile,
        "email": email,
        "game_purchased": game,
        "purchase_date": purchase_date,
        "purchase_time": purchase_time,
        "quantity": qty
    }

def apply_styling(ws):
    bold_font = Font(bold=True)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
    
    ws.freeze_panes = "A2"
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def export_to_excel(file_path, items):
    if not items:
        return
        
    items_list = list(items) if not isinstance(items, dict) else [items]
    
    new_orders = []
    for item in items_list:
        new_orders.append(extract_order_from_details(item))
        
    try:
        # Load or create workbook
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            
        # Ensure correct sheets exist and rename/clean old sheets if present
        sheet_names = wb.sheetnames
        if "Customer Orders" in sheet_names:
            ws_orders = wb["Customer Orders"]
        elif "LTG Customers" in sheet_names:
            ws_orders = wb["LTG Customers"]
            ws_orders.title = "Customer Orders"
        elif "Customer History" in sheet_names:
            ws_orders = wb["Customer History"]
            ws_orders.title = "Customer Orders"
        else:
            if len(sheet_names) == 1 and sheet_names[0] == "Sheet":
                ws_orders = wb.active
                ws_orders.title = "Customer Orders"
            else:
                ws_orders = wb.create_sheet(title="Customer Orders")
                
        if "Customer Summary" in wb.sheetnames:
            ws_summary = wb["Customer Summary"]
        else:
            ws_summary = wb.create_sheet(title="Customer Summary")
            
        # Remove any other sheets
        for name in list(wb.sheetnames):
            if name not in ("Customer Orders", "Customer Summary"):
                del wb[name]
                
        # Get headers of ws_orders
        headers = []
        for cell in ws_orders[1]:
            if cell.value is not None:
                headers.append(str(cell.value))
                
        is_new_format = (headers == ["First Name", "Last Name", "Mobile Number", "Email", "Game Purchased", "Purchase Date", "Purchase Time", "Quantity"])
        
        existing_orders = []
        if is_new_format:
            for row in ws_orders.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_vals = list(row) + [""] * (8 - len(row))
                existing_orders.append({
                    "first_name": str(row_vals[0] or "").strip(),
                    "last_name": str(row_vals[1] or "").strip(),
                    "mobile_number": str(row_vals[2] or "").strip(),
                    "email": str(row_vals[3] or "").strip(),
                    "game_purchased": str(row_vals[4] or "").strip(),
                    "purchase_date": str(row_vals[5] or "").strip(),
                    "purchase_time": str(row_vals[6] or "").strip(),
                    "quantity": int(row_vals[7]) if row_vals[7] is not None else 1
                })
        elif headers:
            # Migrate old format
            for row in ws_orders.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = val
                migrated = extract_order_from_details(row_dict)
                existing_orders.append(migrated)
                
        # Clear ws_orders and ws_summary
        ws_orders.delete_rows(1, ws_orders.max_row + 1)
        ws_summary.delete_rows(1, ws_summary.max_row + 1)
        
        # Write headers
        headers1 = ["First Name", "Last Name", "Mobile Number", "Email", "Game Purchased", "Purchase Date", "Purchase Time", "Quantity"]
        headers2 = ["First Name", "Last Name", "Mobile Number", "Email", "Total Tickets Bought", "Total Orders"]
        
        for col_idx, h in enumerate(headers1, start=1):
            ws_orders.cell(row=1, column=col_idx, value=h)
        for col_idx, h in enumerate(headers2, start=1):
            ws_summary.cell(row=1, column=col_idx, value=h)
            
        # Add new orders to existing orders list if not duplicate
        for no in new_orders:
            is_dup = False
            for eo in existing_orders:
                if (eo["first_name"].lower() == no["first_name"].lower() and
                    eo["last_name"].lower() == no["last_name"].lower() and
                    eo["mobile_number"].strip() == no["mobile_number"].strip() and
                    eo["email"].lower() == no["email"].lower() and
                    eo["game_purchased"].lower() == no["game_purchased"].lower() and
                    eo["purchase_date"] == no["purchase_date"] and
                    eo["purchase_time"] == no["purchase_time"] and
                    eo["quantity"] == no["quantity"]):
                    is_dup = True
                    break
            if not is_dup:
                existing_orders.append(no)
                
        # Write all orders to ws_orders
        for o in existing_orders:
            ws_orders.append([
                o["first_name"],
                o["last_name"],
                o["mobile_number"],
                o["email"],
                o["game_purchased"],
                o["purchase_date"],
                o["purchase_time"],
                o["quantity"]
            ])
            
        # Group by Mobile Number to populate Customer Summary
        grouped = {}
        for o in existing_orders:
            mobile = o["mobile_number"].strip()
            if mobile:
                key = f"mobile:{mobile.lower()}"
            else:
                key = f"composite:{o['first_name'].lower().strip()}|{o['last_name'].lower().strip()}|{o['email'].lower().strip()}"
                
            if key not in grouped:
                grouped[key] = {
                    "first_name": o["first_name"],
                    "last_name": o["last_name"],
                    "mobile_number": o["mobile_number"],
                    "email": o["email"],
                    "total_tickets_bought": 0,
                    "total_orders": 0
                }
            else:
                if grouped[key]["email"] == "N/A" and o["email"] != "N/A":
                    grouped[key]["email"] = o["email"]
                if not grouped[key]["first_name"] and o["first_name"]:
                    grouped[key]["first_name"] = o["first_name"]
                if not grouped[key]["last_name"] and o["last_name"]:
                    grouped[key]["last_name"] = o["last_name"]
                    
            grouped[key]["total_tickets_bought"] += o["quantity"]
            grouped[key]["total_orders"] += 1
            
        for key, cust in grouped.items():
            ws_summary.append([
                cust["first_name"],
                cust["last_name"],
                cust["mobile_number"],
                cust["email"],
                cust["total_tickets_bought"],
                cust["total_orders"]
            ])
            
        # Apply styling to both sheets
        for ws in (ws_orders, ws_summary):
            apply_styling(ws)
            
        wb.save(file_path)
        wb.close()
        print(f"Excel Exporter: Exported successfully to {file_path}")
    except Exception as e:
        print(f"Excel Exporter Error: Failed to export to {file_path}: {e}")

def export_customer_details(details):
    """
    Saves extracted customer/order details into liveticketgroup_customers.xlsx.
    Ensures 2-sheet structure: Customer Orders and Customer Summary.
    """
    export_to_excel(EXCEL_FILE, details)

def export_customer_history(customers):
    """
    Saves all customer history into liveticketgroup_customer_history.xlsx.
    Ensures 2-sheet structure: Customer Orders and Customer Summary.
    """
    export_to_excel(EXCEL_HISTORY_FILE, customers)

def create_styled_workbook(sheets_data):
    """
    sheets_data: list of dicts:
    [
        {
            "title": "Sheet Title",
            "headers": ["Col1", "Col2", ...],
            "rows": [ [val1, val2, ...], ... ]
        },
        ...
    ]
    Returns an openpyxl.Workbook
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    for sheet_info in sheets_data:
        ws = wb.create_sheet(title=sheet_info["title"])
        headers = sheet_info["headers"]
        rows = sheet_info["rows"]
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
        cell_font = Font(name="Segoe UI", size=10)
        border_side = Side(border_style="thin", color="CED6E0")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r in rows:
            ws.append(r)
            
        ws.freeze_panes = "A2"
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = cell_border
                
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    return wb

def workbook_to_bytes(wb):
    import io
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
