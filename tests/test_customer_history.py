import os
import openpyxl
from services.excel_exporter import export_customer_history, EXCEL_HISTORY_FILE

def test_history_export():
    print("Testing Customer History Excel Exporter...")
    
    # 1. Clean existing file
    if os.path.exists(EXCEL_HISTORY_FILE):
        os.remove(EXCEL_HISTORY_FILE)
        print("Removed existing test Excel history file.")
        
    # 2. Mock customers list
    mock_customers = [
        {
            "first_name": "Alice",
            "last_name": "Smith",
            "phone_number": "+447123456789",
            "email": "alice@example.com"
        },
        {
            "first_name": "Bob",
            "last_name": "Jones",
            "phone_number": "", # Missing phone, duplicate check will use composite
            "email": "bob@example.com"
        },
        {
            "first_name": "Charlie",
            "last_name": "Brown",
            "phone_number": "+447999888777",
            "email": "" # Empty email, should default to N/A
        }
    ]
    
    print("\nExporting initial list of customers...")
    export_customer_history(mock_customers)
    
    # Verify file exists
    if os.path.exists(EXCEL_HISTORY_FILE):
        print("Success: Excel history file created.")
    else:
        print("Error: Excel history file was not created.")
        return
        
    # 3. Test duplicates
    test_duplicates = [
        {
            "first_name": "Alice",
            "last_name": "Smith",
            "phone_number": "+447123456789", # Duplicate phone
            "email": "anotheremail@example.com"
        },
        {
            "first_name": "Bob",
            "last_name": "Jones",
            "phone_number": "", # Missing phone, duplicate name + name + email
            "email": "bob@example.com"
        },
        {
            "first_name": "David",
            "last_name": "Miller",
            "phone_number": "+447555666777", # Unique
            "email": "david@example.com"
        }
    ]
    
    print("\nExporting duplicates and one new customer...")
    export_customer_history(test_duplicates)
    
    # 4. Verify results
    wb = openpyxl.load_workbook(EXCEL_HISTORY_FILE)
    print("\nVerifying sheets in workbook:")
    print("Sheet names:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet '{sheet_name}' properties:")
        print("Max Row:", ws.max_row)
        print("Max Column:", ws.max_column)
        
        headers = [cell.value for cell in ws[1]]
        print("Headers:", headers)
        
        # Check bold headers
        all_headers_bold = all(ws.cell(row=1, column=c).font.bold for c in range(1, len(headers) + 1))
        print("Are headers bold?", all_headers_bold)
        print("Freeze panes:", ws.freeze_panes)
        
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            print(f"Row {r}:", row_vals)
            
    wb.close()
    
    # Delete temporary test history file
    if os.path.exists(EXCEL_HISTORY_FILE):
        os.remove(EXCEL_HISTORY_FILE)
        print("\nCleaned up test Excel file.")

def test_export_script_helpers():
    from services.excel_exporter import split_customer_name, format_purchase_datetime, extract_order_from_details
    from datetime import datetime
    
    assert split_customer_name("John Doe") == ("John", "Doe")
    assert split_customer_name("Alice") == ("Alice", "")
    assert split_customer_name("") == ("", "")
    assert split_customer_name("-") == ("", "")
    
    dt = datetime(2025, 1, 10, 14, 30, 45)
    assert format_purchase_datetime(dt) == ("2025-01-10", "14:30:45")
    
    assert format_purchase_datetime("10-01-2025 14:30:45") == ("2025-01-10", "14:30:45")
    assert format_purchase_datetime("2025-01-10T14:30:45Z") == ("2025-01-10", "14:30:45")
    assert format_purchase_datetime("Invalid Date") == ("Invalid", "Date")
    assert format_purchase_datetime(None) == ("N/A", "N/A")
    
    mock_details = {
        "billing_full_name": "Charlie Brown",
        "mobile_number": "+447777888999",
        "email": "charlie@peanuts.com",
        "event_name": "Snoopy Show",
        "sale_date": "10-01-2025 14:30:00",
        "quantity": 3
    }
    info = extract_order_from_details(mock_details)
    assert info["first_name"] == "Charlie"
    assert info["last_name"] == "Brown"
    assert info["mobile_number"] == "+447777888999"
    assert info["email"] == "charlie@peanuts.com"
    assert info["game_purchased"] == "Snoopy Show"
    assert info["purchase_date"] == "2025-01-10"
    assert info["purchase_time"] == "14:30:00"
    assert info["quantity"] == 3

if __name__ == "__main__":
    test_history_export()
    test_export_script_helpers()
