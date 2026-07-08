import os
import shutil
import openpyxl
from services.excel_exporter import export_customer_details, EXCEL_FILE

def test_excel_export():
    print("Testing Excel Exporter...")
    
    # Remove existing excel file if it exists for a clean test
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
        print("Removed existing test Excel file.")

    # 1. Export first dummy order details
    order1 = {
        "order_number": "111111",
        "event_name": "Arsenal vs Chelsea",
        "customer_name": "Alice Smith",
        "quantity": 2,
        "total_amount": 250.00,
        "currency": "£"
    }
    
    print("\nExporting first order...")
    export_customer_details(order1)
    
    # Verify file is created
    if os.path.exists(EXCEL_FILE):
        print("Success: Excel file created.")
    else:
        print("Error: Excel file was not created.")
        return

    # 2. Try exporting duplicate order
    print("\nExporting duplicate order (should skip)...")
    export_customer_details(order1)
    
    # 3. Export second order with a new field/column
    order2 = {
        "order_number": "222222",
        "event_name": "Man Utd vs Man City",
        "customer_name": "Bob Jones",
        "quantity": 4,
        "total_amount": 600.00,
        "currency": "£",
        "new_platform_field": "Special Value" # Dynamic column - should be ignored!
    }
    
    print("\nExporting second order with dynamic field (ignored)...")
    export_customer_details(order2)
    
    # 4. Verify sheet content using openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    print("\nVerifying sheets in workbook:")
    print("Sheet names:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet '{sheet_name}' properties:")
        print("Max Row:", ws.max_row)
        print("Max Column:", ws.max_column)
        
        headers = [cell.value for cell in ws[1]]
        print("Headers:", headers)
        
        # Check bold headers
        all_headers_bold = all(ws.cell(row=1, column=c).font.bold for c in range(1, len(headers) + 1))
        print("Are headers bold?", all_headers_bold)
        print("Freeze panes:", ws.freeze_panes)
        
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
            print(f"Row {r}:", row_vals)
            
    wb.close()
    
    # Clean up test file
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
        print("\nCleaned up test Excel file.")

if __name__ == "__main__":
    test_excel_export()
