import os
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Add project root directory to Python path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.append(project_root)

from platforms.footballticketnet.customer_history_scraper import scrape_ftn_customer_history

def main():
    try:
        orders = scrape_ftn_customer_history()
        
        # 1. Group customers for Customer Summary
        grouped_customers = {}
        for o in orders:
            mobile = o["mobile_number"].strip()
            if mobile and mobile.lower() != "n/a":
                key = f"mobile:{mobile.lower()}"
            else:
                key = f"composite:{o['first_name'].lower().strip()}|{o['last_name'].lower().strip()}|{o['email'].lower().strip()}"
                
            if key not in grouped_customers:
                grouped_customers[key] = {
                    "first_name": o["first_name"],
                    "last_name": o["last_name"],
                    "mobile_number": o["mobile_number"],
                    "email": o["email"],
                    "total_tickets_bought": 0,
                    "total_orders": 0
                }
            else:
                if grouped_customers[key]["email"] == "N/A" and o["email"] != "N/A":
                    grouped_customers[key]["email"] = o["email"]
                if not grouped_customers[key]["first_name"] and o["first_name"]:
                    grouped_customers[key]["first_name"] = o["first_name"]
                if not grouped_customers[key]["last_name"] and o["last_name"]:
                    grouped_customers[key]["last_name"] = o["last_name"]
                    
            grouped_customers[key]["total_tickets_bought"] += o["quantity"]
            grouped_customers[key]["total_orders"] += 1
            
        excel_file_path = os.path.join(project_root, "footballticketnet_customer_history.xlsx")
        
        wb = openpyxl.Workbook()
        
        # 2. Saving Customer Orders sheet
        print("Saving Customer Orders sheet...")
        ws1 = wb.active
        ws1.title = "Customer Orders"
        
        headers1 = ["First Name", "Last Name", "Mobile Number", "Email", "Game Purchased", "Purchase Date", "Purchase Time", "Quantity"]
        ws1.append(headers1)
        
        seen_orders = set()
        for o in orders:
            oid = o.get("order_id")
            if oid and not oid.startswith("FTN_TEMP_"):
                if oid in seen_orders:
                    continue
                seen_orders.add(oid)
            else:
                comp = (
                    o["first_name"].lower(),
                    o["last_name"].lower(),
                    o["mobile_number"].strip(),
                    o["email"].lower(),
                    o["game_purchased"].lower(),
                    o["purchase_date"],
                    o["purchase_time"],
                    o["quantity"]
                )
                if comp in seen_orders:
                    continue
                seen_orders.add(comp)
                
            ws1.append([
                o["first_name"],
                o["last_name"],
                o["mobile_number"],
                o["email"],
                o["game_purchased"],
                o["purchase_date"],
                o["purchase_time"],
                o["quantity"]
            ])
            
        # 3. Saving Customer Summary sheet
        print("Saving Customer Summary sheet...")
        ws2 = wb.create_sheet(title="Customer Summary")
        
        headers2 = ["First Name", "Last Name", "Mobile Number", "Email", "Total Tickets Bought", "Total Orders"]
        ws2.append(headers2)
        
        for key, cust in grouped_customers.items():
            ws2.append([
                cust["first_name"],
                cust["last_name"],
                cust["mobile_number"],
                cust["email"],
                cust["total_tickets_bought"],
                cust["total_orders"]
            ])
            
        bold_font = Font(bold=True)
        for ws in (ws1, ws2):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
            ws.freeze_panes = "A2"
            
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    lines = val_str.split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(excel_file_path)
        wb.close()
        
        print("Export finished.")
        
        # Log failed orders if any exist
        failed_log = os.path.join(project_root, "ftn_failed_orders.log")
        if os.path.exists(failed_log):
            print("\nSome orders failed to scrape. Check references in ftn_failed_orders.log:")
            with open(failed_log, "r", encoding="utf-8") as f:
                print(f.read().strip())
                
    except Exception as e:
        print(f"Error during export: {e}")

if __name__ == "__main__":
    main()
